"""
Integration tests for ACC (Autodesk Construction Cloud) Excel export.

Verifies the full export pipeline produces a valid ACC checklist template:
- 22-column structure (A–V) matching the reference format
- Lookups sheet with valid enumeration values
- Correct response type mapping (slash-delimited, matching Lookups)
- ACC-convention template naming ({spec}_{level}_{equipment}_{activity})
- Template type mapping (Quality / Commissioning)
- Section rows use Response Type = "Section"
- Non-conforming answers and Issue Auto Create are conditional
- Section descriptions and priority metadata are preserved
- Backward-compat aliases (export_to_procore_excel, export_to_excel)

Run with:  pytest itc_form_generator/test_acc_export.py -v
"""

import io
import pytest
from openpyxl import load_workbook

from itc_form_generator.models import (
    InspectionForm, FormSection, CheckItem,
    FormType, CheckItemType, Priority,
)
from itc_form_generator.exporter import (
    FormExporter, ACC_HEADERS, escape_excel_formula,
)


# ── Fixtures ────────────────────────────────────────────────────────


@pytest.fixture
def exporter():
    return FormExporter()


@pytest.fixture
def pfi_form():
    """A Pre-Functional Inspection form with two sections and mixed check types."""
    return InspectionForm(
        form_type=FormType.PFI,
        title="Jockey Pump Pre-Functional Inspection",
        system="Jockey Pump",
        system_tag="JP-01",
        project="213113",
        sections=[
            FormSection(
                title="Physical Installation",
                description="Verify physical installation of jockey pump and piping.",
                check_items=[
                    CheckItem(
                        id="PI-001",
                        description="Verify pump is anchored to housekeeping pad",
                        check_type=CheckItemType.VISUAL,
                        priority=Priority.CRITICAL,
                        acceptance_criteria="Pump bolted per manufacturer specs",
                        method="Visual inspection",
                        expected_value="4 anchor bolts, torqued to spec",
                    ),
                    CheckItem(
                        id="PI-002",
                        description="Record suction pressure gauge reading",
                        check_type=CheckItemType.MEASUREMENT,
                        priority=Priority.HIGH,
                        acceptance_criteria="Within design range",
                        expected_value="40-60 PSI",
                    ),
                ],
            ),
            FormSection(
                title="Electrical Connections",
                description="Verify electrical power and control wiring.",
                check_items=[
                    CheckItem(
                        id="EC-001",
                        description="Verify motor rotation direction",
                        check_type=CheckItemType.FUNCTIONAL,
                        priority=Priority.CRITICAL,
                        acceptance_criteria="Clockwise when viewed from drive end",
                    ),
                    CheckItem(
                        id="EC-002",
                        description="Confirm nameplate data matches submittals",
                        check_type=CheckItemType.DOCUMENTATION,
                        priority=Priority.MEDIUM,
                    ),
                    CheckItem(
                        id="EC-003",
                        description="Verify ground fault protection is installed",
                        check_type=CheckItemType.VERIFICATION,
                        priority=Priority.LOW,
                    ),
                ],
            ),
        ],
    )


@pytest.fixture
def fpt_form():
    """A Functional Performance Test form."""
    return InspectionForm(
        form_type=FormType.FPT,
        title="CRAH Functional Performance Test",
        system="CRAH Unit",
        system_tag="CRAH-01",
        project="",
        sections=[
            FormSection(
                title="Cooling Mode Test",
                check_items=[
                    CheckItem(
                        id="CM-001",
                        description="Verify supply air temperature control",
                        check_type=CheckItemType.FUNCTIONAL,
                        priority=Priority.HIGH,
                        acceptance_criteria="±2°F of setpoint",
                        expected_value="55°F",
                    ),
                ],
            ),
        ],
    )


@pytest.fixture
def pfi_workbook(exporter, pfi_form):
    """Export pfi_form and return a loaded openpyxl Workbook."""
    excel_bytes = exporter.export_to_acc_excel([pfi_form], project_number="213113")
    assert excel_bytes is not None
    return load_workbook(io.BytesIO(excel_bytes))


@pytest.fixture
def multi_workbook(exporter, pfi_form, fpt_form):
    """Export two forms and return a loaded openpyxl Workbook."""
    excel_bytes = exporter.export_to_acc_excel(
        [pfi_form, fpt_form], project_number="213113"
    )
    assert excel_bytes is not None
    return load_workbook(io.BytesIO(excel_bytes))


# ── Helpers ─────────────────────────────────────────────────────────


def _row_values(ws, row_num, max_col=22):
    """Return a list of cell values for a row (1-indexed columns)."""
    return [ws.cell(row=row_num, column=c).value for c in range(1, max_col + 1)]


def _col_values(ws, col_num, start_row=4):
    """Return all non-None values in a column starting from start_row."""
    vals = []
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col_num).value
        if v is not None:
            vals.append(v)
    return vals


# ═══════════════════════════════════════════════════════════════════
#  Test classes
# ═══════════════════════════════════════════════════════════════════


class TestWorkbookStructure:
    """Verify workbook-level structure: sheets, header rows, column count."""

    def test_single_form_produces_three_sheets(self, pfi_workbook):
        names = pfi_workbook.sheetnames
        assert "FULL" in names
        assert "Lookups" in names
        assert len(names) == 3  # FULL + per-form sheet + Lookups

    def test_multi_form_produces_correct_sheets(self, multi_workbook):
        names = multi_workbook.sheetnames
        assert "FULL" in names
        assert "Lookups" in names
        assert len(names) == 4  # FULL + 2 per-form sheets + Lookups

    def test_full_sheet_is_first(self, pfi_workbook):
        assert pfi_workbook.sheetnames[0] == "FULL"

    def test_lookups_sheet_is_last(self, pfi_workbook):
        assert pfi_workbook.sheetnames[-1] == "Lookups"

    def test_per_form_sheet_name_truncated_to_31(self, exporter):
        form = InspectionForm(
            form_type=FormType.PFI,
            title="Test",
            system="VeryLongSystemName",
            system_tag="ABCDEFGHIJKLMNOPQRSTUVWXYZ_TAG",
            sections=[],
        )
        wb_bytes = exporter.export_to_acc_excel([form])
        wb = load_workbook(io.BytesIO(wb_bytes))
        for name in wb.sheetnames:
            assert len(name) <= 31


class TestHeaderRows:
    """Verify rows 1-3 match the ACC template specification."""

    def test_row1_has_22_acc_headers(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        headers = _row_values(ws, 1)
        assert headers == ACC_HEADERS

    def test_row2_has_22_instructions(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        row2 = _row_values(ws, 2)
        assert all(v is not None for v in row2), "Every instruction cell should be populated"
        assert len(row2) == 22

    def test_row3_is_reminder(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        assert ws.cell(row=3, column=1).value == "###"
        assert "Don't delete" in str(ws.cell(row=3, column=2).value)

    def test_per_form_sheet_also_has_headers(self, pfi_workbook):
        per_form_sheet = [s for s in pfi_workbook.sheetnames if s not in ("FULL", "Lookups")][0]
        ws = pfi_workbook[per_form_sheet]
        headers = _row_values(ws, 1)
        assert headers == ACC_HEADERS

    def test_freeze_panes_at_a4(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        assert ws.freeze_panes == "A4"


class TestResponseTypeMapping:
    """Verify check-type → response-type mapping matches Lookups values."""

    VALID_RESPONSE_TYPES = {
        "Section", "Yes/No/N/A", "Plus/Minus/N/A",
        "True/False/N/A", "Pass/Fail/N/A", "Text", "Date",
    }

    def test_visual_maps_to_yes_no_na(self, exporter):
        assert exporter._get_response_type(CheckItemType.VISUAL) == "Yes/No/N/A"

    def test_measurement_maps_to_text(self, exporter):
        assert exporter._get_response_type(CheckItemType.MEASUREMENT) == "Text"

    def test_functional_maps_to_pass_fail_na(self, exporter):
        assert exporter._get_response_type(CheckItemType.FUNCTIONAL) == "Pass/Fail/N/A"

    def test_documentation_maps_to_text(self, exporter):
        assert exporter._get_response_type(CheckItemType.DOCUMENTATION) == "Text"

    def test_verification_maps_to_yes_no_na(self, exporter):
        assert exporter._get_response_type(CheckItemType.VERIFICATION) == "Yes/No/N/A"

    def test_all_data_rows_use_valid_response_types(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        col_f = _col_values(ws, 6, start_row=4)  # Column F = Response Type
        for val in col_f:
            assert val in self.VALID_RESPONSE_TYPES, f"Invalid response type: {val}"


class TestNonConformingAnswers:
    """Verify non-conforming answer logic."""

    def test_yes_no_na_non_conforming_is_no(self, exporter):
        assert exporter._get_non_conforming("Yes/No/N/A") == "No"

    def test_pass_fail_na_non_conforming_is_fail(self, exporter):
        assert exporter._get_non_conforming("Pass/Fail/N/A") == "Fail"

    def test_text_has_no_non_conforming(self, exporter):
        assert exporter._get_non_conforming("Text") == ""

    def test_section_has_no_non_conforming(self, exporter):
        assert exporter._get_non_conforming("Section") == ""

    def test_non_conforming_column_in_workbook(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            resp = ws.cell(row=r, column=6).value  # F: Response Type
            nc = ws.cell(row=r, column=9).value     # I: Non Conforming Answers
            if resp == "Yes/No/N/A":
                assert nc == "No", f"Row {r}: Yes/No/N/A should have NC='No'"
            elif resp == "Pass/Fail/N/A":
                assert nc == "Fail", f"Row {r}: Pass/Fail/N/A should have NC='Fail'"
            elif resp in ("Text", "Section"):
                assert nc is None, f"Row {r}: {resp} should have no NC value"


class TestIssueAutoCreate:
    """Issue Auto Create should only be TRUE when non-conforming is set."""

    def test_auto_create_only_when_non_conforming(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            resp = ws.cell(row=r, column=6).value   # F: Response Type
            auto = ws.cell(row=r, column=19).value   # S: Issue Auto Create
            if resp in ("Text", "Section"):
                assert auto is None, f"Row {r}: {resp} should not auto-create issues"
            elif resp in ("Yes/No/N/A", "Pass/Fail/N/A"):
                assert auto == "TRUE", f"Row {r}: {resp} should auto-create issues"


class TestTemplateNaming:
    """Verify ACC-convention template name: {spec}_{level}_{equipment}_{activity}."""

    def test_pfi_template_name(self, exporter, pfi_form):
        name = exporter._get_template_name(pfi_form, project_number="213113")
        assert name == "213113_L2_JP-01_SetInPlace"

    def test_fpt_template_name(self, exporter, fpt_form):
        name = exporter._get_template_name(fpt_form, project_number="999999")
        assert name == "999999_L3_CRAH-01_FunctionalTest"

    def test_fallback_project_number(self, exporter):
        form = InspectionForm(
            form_type=FormType.IST, title="T", system="Sys",
            system_tag="TAG", project="PROJ42", sections=[],
        )
        name = exporter._get_template_name(form, project_number="")
        assert name.startswith("PROJ42_")

    def test_fallback_to_default_project(self, exporter):
        form = InspectionForm(
            form_type=FormType.IST, title="T", system="Sys",
            system_tag="TAG", project="", sections=[],
        )
        name = exporter._get_template_name(form, project_number="")
        assert name.startswith("000000_")

    def test_fallback_equipment_from_system(self, exporter):
        form = InspectionForm(
            form_type=FormType.PFI, title="T", system="Air Handler Unit",
            system_tag="", sections=[],
        )
        name = exporter._get_template_name(form, project_number="123")
        assert "AirHandlerUnit" in name

    def test_template_name_on_every_data_row(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            val = ws.cell(row=r, column=2).value  # B: Template Name
            assert val is not None, f"Row {r}: Template Name must not be empty"
            assert "_L2_" in val


class TestTemplateType:
    """Verify FormType → ACC Template Type mapping."""

    def test_pfi_is_quality(self, exporter):
        form = InspectionForm(
            form_type=FormType.PFI, title="T", system="S", sections=[],
        )
        assert exporter._get_template_type(form) == "Quality"

    def test_fpt_is_commissioning(self, exporter):
        form = InspectionForm(
            form_type=FormType.FPT, title="T", system="S", sections=[],
        )
        assert exporter._get_template_type(form) == "Commissioning"

    def test_ist_is_commissioning(self, exporter):
        form = InspectionForm(
            form_type=FormType.IST, title="T", system="S", sections=[],
        )
        assert exporter._get_template_type(form) == "Commissioning"

    def test_template_type_on_every_data_row(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        valid_types = {"Quality", "Safety", "Punch List", "Commissioning"}
        for r in range(4, ws.max_row + 1):
            val = ws.cell(row=r, column=3).value  # C: Template Type
            assert val in valid_types, f"Row {r}: unexpected template type '{val}'"


class TestSectionRows:
    """Verify section header rows use Response Type = 'Section'."""

    def test_section_rows_have_response_type_section(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        section_rows = [
            r for r in range(4, ws.max_row + 1)
            if ws.cell(row=r, column=6).value == "Section"
        ]
        assert len(section_rows) == 2  # "Physical Installation" + "Electrical Connections"

    def test_section_title_in_item_text(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        section_titles = []
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=6).value == "Section":
                section_titles.append(ws.cell(row=r, column=4).value)
        assert "Physical Installation" in section_titles
        assert "Electrical Connections" in section_titles

    def test_section_description_preserved(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            if ws.cell(row=r, column=4).value == "Physical Installation":
                desc = ws.cell(row=r, column=5).value  # E: Item Description
                assert desc is not None
                assert "jockey pump" in desc.lower()
                break
        else:
            pytest.fail("Section 'Physical Installation' not found")


class TestItemRows:
    """Verify check-item data rows populate the correct columns."""

    def test_total_item_rows(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        item_rows = [
            r for r in range(4, ws.max_row + 1)
            if ws.cell(row=r, column=6).value not in ("Section", None)
        ]
        assert len(item_rows) == 5  # 2 + 3 check items

    def test_item_text_populated(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            resp = ws.cell(row=r, column=6).value
            if resp and resp != "Section":
                assert ws.cell(row=r, column=4).value  # D: Item Text

    def test_item_description_includes_acceptance_criteria(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        # PI-001 has acceptance_criteria = "Pump bolted per manufacturer specs"
        found = False
        for r in range(4, ws.max_row + 1):
            text = ws.cell(row=r, column=4).value or ""
            if "anchored to housekeeping pad" in text:
                desc = ws.cell(row=r, column=5).value or ""
                assert "Pump bolted" in desc
                assert "Method: Visual inspection" in desc
                assert "Expected: 4 anchor bolts" in desc
                found = True
                break
        assert found, "PI-001 item not found"

    def test_critical_priority_in_description(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            text = ws.cell(row=r, column=4).value or ""
            if "anchored to housekeeping pad" in text:
                desc = ws.cell(row=r, column=5).value or ""
                assert "Priority: Critical" in desc
                break

    def test_medium_priority_omitted_from_description(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            text = ws.cell(row=r, column=4).value or ""
            if "nameplate data" in text:
                desc = ws.cell(row=r, column=5).value or ""
                assert "Priority:" not in desc
                break

    def test_response_required_critical_high(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            text = ws.cell(row=r, column=4).value or ""
            required = ws.cell(row=r, column=7).value  # G
            if "anchored to housekeeping pad" in text:
                assert required == "TRUE"  # CRITICAL
            elif "suction pressure gauge" in text:
                assert required == "TRUE"  # HIGH
            elif "nameplate data" in text:
                assert required == "FALSE"  # MEDIUM
            elif "ground fault protection" in text:
                assert required == "FALSE"  # LOW

    def test_issue_type_matches_template_type(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            resp = ws.cell(row=r, column=6).value
            if resp and resp != "Section":
                issue_type = ws.cell(row=r, column=12).value  # L
                template_type = ws.cell(row=r, column=3).value  # C
                assert issue_type == template_type

    def test_issue_subtype_is_inspection(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(4, ws.max_row + 1):
            resp = ws.cell(row=r, column=6).value
            if resp and resp != "Section":
                assert ws.cell(row=r, column=13).value == "Inspection"  # M

    def test_index_column_populated(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        indices = _col_values(ws, 20, start_row=4)  # T: index
        assert len(indices) > 0
        assert all(isinstance(i, int) for i in indices)


class TestLookupsSheet:
    """Verify the Lookups sheet contains valid ACC enumeration values."""

    def test_lookups_version(self, pfi_workbook):
        ws = pfi_workbook["Lookups"]
        assert ws.cell(row=1, column=1).value == "Version"
        assert ws.cell(row=1, column=2).value == 1

    def test_lookups_template_types(self, pfi_workbook):
        ws = pfi_workbook["Lookups"]
        assert ws.cell(row=3, column=1).value == "Template Types"
        types = [ws.cell(row=r, column=1).value for r in range(4, 8)]
        assert types == ["Quality", "Safety", "Punch List", "Commissioning"]

    def test_lookups_response_types(self, pfi_workbook):
        ws = pfi_workbook["Lookups"]
        assert ws.cell(row=9, column=1).value == "Response Types"
        rtypes = [ws.cell(row=r, column=1).value for r in range(10, 17)]
        assert rtypes == [
            "Section", "Yes/No/N/A", "Plus/Minus/N/A",
            "True/False/N/A", "Pass/Fail/N/A", "Text", "Date",
        ]


class TestMultiFormExport:
    """Verify multi-form export produces correct FULL sheet and per-form sheets."""

    def test_full_sheet_contains_all_items(self, multi_workbook, pfi_form, fpt_form):
        ws = multi_workbook["FULL"]
        total_sections = len(pfi_form.sections) + len(fpt_form.sections)
        total_items = pfi_form.total_items + fpt_form.total_items
        expected_data_rows = total_sections + total_items

        data_rows = [
            r for r in range(4, ws.max_row + 1)
            if ws.cell(row=r, column=2).value is not None
        ]
        assert len(data_rows) == expected_data_rows

    def test_full_sheet_has_both_template_names(self, multi_workbook):
        ws = multi_workbook["FULL"]
        template_names = set(_col_values(ws, 2, start_row=4))
        assert len(template_names) == 2
        assert any("SetInPlace" in n for n in template_names)
        assert any("FunctionalTest" in n for n in template_names)

    def test_per_form_sheets_have_correct_counts(self, multi_workbook, pfi_form, fpt_form):
        sheets = [s for s in multi_workbook.sheetnames if s not in ("FULL", "Lookups")]
        assert len(sheets) == 2

        for sheet_name in sheets:
            ws = multi_workbook[sheet_name]
            rows = [
                r for r in range(4, ws.max_row + 1)
                if ws.cell(row=r, column=2).value is not None
            ]
            assert len(rows) > 0  # each sheet has data


class TestBackwardCompatibility:
    """Verify old method names still work."""

    def test_export_to_procore_excel_alias(self, exporter, pfi_form):
        result = exporter.export_to_procore_excel([pfi_form], project_number="123")
        assert result is not None
        assert isinstance(result, bytes)

    def test_export_to_excel_wrapper(self, exporter, pfi_form):
        result = exporter.export_to_excel([pfi_form], project_number="123")
        assert result is not None
        assert isinstance(result, bytes)

    def test_all_three_methods_produce_identical_output(self, exporter, pfi_form):
        a = exporter.export_to_acc_excel([pfi_form], project_number="X")
        b = exporter.export_to_procore_excel([pfi_form], project_number="X")
        c = exporter.export_to_excel([pfi_form], project_number="X")
        assert a == b == c


class TestEdgeCases:
    """Edge cases and defensive behavior."""

    def test_empty_forms_list(self, exporter):
        result = exporter.export_to_acc_excel([], project_number="123")
        assert result is not None
        wb = load_workbook(io.BytesIO(result))
        assert "FULL" in wb.sheetnames
        assert "Lookups" in wb.sheetnames

    def test_form_with_no_sections(self, exporter):
        form = InspectionForm(
            form_type=FormType.CXC, title="Empty", system="Sys",
            system_tag="SYS-01", sections=[],
        )
        result = exporter.export_to_acc_excel([form], project_number="456")
        assert result is not None
        wb = load_workbook(io.BytesIO(result))
        ws = wb["FULL"]
        data_rows = [r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=2).value]
        assert len(data_rows) == 0

    def test_section_with_no_items(self, exporter):
        form = InspectionForm(
            form_type=FormType.PFI, title="T", system="S",
            system_tag="S-01",
            sections=[FormSection(title="Empty Section", check_items=[])],
        )
        result = exporter.export_to_acc_excel([form])
        wb = load_workbook(io.BytesIO(result))
        ws = wb["FULL"]
        data_rows = [r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=2).value]
        assert len(data_rows) == 1  # just the section row
        assert ws.cell(row=4, column=6).value == "Section"

    def test_formula_injection_escaped(self, exporter):
        form = InspectionForm(
            form_type=FormType.PFI, title="T", system="S",
            system_tag="S-01",
            sections=[
                FormSection(
                    title="=HYPERLINK(\"evil\")",
                    check_items=[
                        CheckItem(
                            id="1",
                            description="+cmd('calc')",
                            check_type=CheckItemType.VISUAL,
                        ),
                    ],
                ),
            ],
        )
        result = exporter.export_to_acc_excel([form])
        wb = load_workbook(io.BytesIO(result))
        ws = wb["FULL"]
        section_text = ws.cell(row=4, column=4).value
        assert section_text.startswith("'"), "Formula should be escaped with single quote"
        item_text = ws.cell(row=5, column=4).value
        assert item_text.startswith("'"), "Formula should be escaped with single quote"

    def test_escape_excel_formula_function(self):
        assert escape_excel_formula("=SUM(A1)") == "'=SUM(A1)"
        assert escape_excel_formula("-1+1") == "'-1+1"
        assert escape_excel_formula("+cmd") == "'+cmd"
        assert escape_excel_formula("@mention") == "'@mention"
        assert escape_excel_formula("Normal text") == "Normal text"
        assert escape_excel_formula("") == ""

    def test_all_form_types_produce_valid_output(self, exporter):
        for ft in FormType:
            form = InspectionForm(
                form_type=ft, title=f"{ft.name} Test",
                system="TestSys", system_tag="TS-01",
                sections=[
                    FormSection(
                        title="Section",
                        check_items=[
                            CheckItem(id="1", description="Item",
                                      check_type=CheckItemType.VISUAL),
                        ],
                    ),
                ],
            )
            result = exporter.export_to_acc_excel([form])
            assert result is not None, f"FormType {ft.name} should export successfully"
            wb = load_workbook(io.BytesIO(result))
            assert "FULL" in wb.sheetnames
            assert "Lookups" in wb.sheetnames


class TestColumnCount:
    """Ensure consistent 22-column structure across all rows."""

    def test_headers_list_has_22_entries(self):
        assert len(ACC_HEADERS) == 22

    def test_row1_has_22_columns(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        row1 = _row_values(ws, 1, max_col=22)
        assert len(row1) == 22
        assert row1[0] == "###"          # A
        assert row1[21] == "required by"  # V

    def test_no_data_beyond_column_v(self, pfi_workbook):
        ws = pfi_workbook["FULL"]
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=23).value
            assert val is None, f"Row {r} has data beyond column V (col 23)"
