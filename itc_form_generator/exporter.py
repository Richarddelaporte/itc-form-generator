"""Export ITC forms to CSV and Excel formats.

Supports ACC (Autodesk Construction Cloud) checklist import format
for BIM360/ACC field inspection workflows.
"""

import csv
import io
from typing import Optional

from .models import InspectionForm, FormSection, CheckItem, FormType, CheckItemType, Priority


def escape_excel_formula(value: str) -> str:
    """Escape strings that Excel might interpret as formulas."""
    if not value:
        return value
    if isinstance(value, str) and value and value[0] in ('=', '-', '+', '@'):
        return "'" + value
    return value


# ACC column headers (A–V, 22 columns)
ACC_HEADERS = [
    '###',                  # A
    'Template Name',        # B
    'Template Type',        # C
    'Item Text',            # D
    'Item Description',     # E
    'Response Type',        # F
    'Response Required',    # G
    'List Answers',         # H
    'Non Conforming Answers',  # I
    'Issue Title',          # J
    'Issue Description',    # K
    'Issue Type',           # L
    'Issue Subtype',        # M
    'Issue Assignee',       # N
    'Issue Assignee Type',  # O
    'Issue Owner',          # P
    'Issue Root Cause Category',  # Q
    'Issue Root Cause',     # R
    'Issue Auto Create',    # S
    'index',                # T
    'number',               # U
    'required by',          # V
]

# Row 2: Required / Optional labels
ACC_ROW2_REQUIRED = [
    '',           # A: ###
    'Required',   # B
    'Required',   # C
    'Required',   # D
    'Optional',   # E
    'Optional',   # F
    'Required',   # G (Response Required)
    '',           # H
    'Optional',   # I
    'Optional',   # J
    '',           # K
    'Optional',   # L
    'Optional',   # M
    'Optional',   # N
    'Optional',   # O
    'Optional',   # P
    'Optional',   # Q
    'Optional',   # R
    'Optional',   # S
    'Optional',   # T
    'Optional',   # U
    'Optional',   # V
]

# Row 3: Short descriptions
ACC_ROW3_DESCRIPTIONS = [
    '',
    'Name of the template.',
    'Type of the template.',
    'Text for the template item OR section name.',
    'Description for the template item OR section.',
    'To create a section, enter Section.\n\nOtherwise, enter one of the below:\nYes, No, N/A\nTrue, False, N/A\nPass, Fail, N/A\nText\nMultiple Choice\nCheckboxes',
    'Controls whether response is required.\n\nEnter TRUE or FALSE.\n\nIf not set, response will be required for all response types except Text.',
    '',
    'For Multiple Choice and Checkbox response types, specify the answers from the List Answers column that are non-conforming.\n\nUse | to separate each answer in the list.',
    '',
    '',
    'Type of the issue.',
    'SubType of the issue.',
    'UserID of the issue assignee.',
    'Enter USER, ROLE, or COMPANY.',
    'UserID of the issue owner.',
    'Category of Root Cause of the Issue.',
    'Root Cause of the Issue.',
    'Controls whether issue needs to be auto created for non-conforming responses.\n\nEnter TRUE or FALSE.',
    'Optional',
    'Optional',
    'Optional',
]

# Row 4: Paste instructions
ACC_ROW4_INSTRUCTIONS = [
    '',
    'Must be filled in next to each template item – use Copy and Paste.',
    'Must be filled in next to each template item – use Copy and Paste.',
    '',
    '',
    '',
    '',
    '',
    'For Multiple Choice and Checkbox response types, specify the answers from the List Answers column that are non-conforming.\n\nUse | to separate each answer in the list.',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
    '',
]

ACC_COLUMN_WIDTHS = [
    6, 50, 18, 80, 50, 20, 16, 20, 30,
    20, 20, 14, 14, 20, 18, 20, 22, 18,
    16, 8, 8, 14,
]

# FormType → ACC Level code
FORM_TYPE_LEVEL = {
    FormType.PFI: 'L2',
    FormType.FPT: 'L3',
    FormType.IST: 'L4',
    FormType.CXC: 'L2',
    FormType.ITC: 'L3',
}

# FormType → ACC activity label
FORM_TYPE_ACTIVITY = {
    FormType.PFI: 'SetInPlace',
    FormType.FPT: 'FunctionalTest',
    FormType.IST: 'Integration',
    FormType.CXC: 'Commissioning',
    FormType.ITC: 'ITC',
}

# FormType → ACC Template Type
FORM_TYPE_TEMPLATE_TYPE = {
    FormType.PFI: 'Quality',
    FormType.FPT: 'Commissioning',
    FormType.IST: 'Commissioning',
    FormType.CXC: 'Commissioning',
    FormType.ITC: 'Commissioning',
}


class FormExporter:
    """Export inspection forms to various formats."""

    # ── CSV (unchanged) ──────────────────────────────────────────────

    def export_to_csv(self, form: InspectionForm) -> str:
        """Export a single form to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'Form Type', 'System', 'System Tag', 'Section', 'Item ID',
            'Description', 'Check Type', 'Priority', 'Acceptance Criteria',
            'Expected Value', 'Actual Value', 'Pass/Fail', 'Comments',
        ])
        for section in form.sections:
            for item in section.check_items:
                writer.writerow([
                    form.form_type.value, form.system, form.system_tag,
                    section.title, item.id, item.description,
                    item.check_type.value, item.priority.value,
                    item.acceptance_criteria, item.expected_value,
                    item.actual_value, item.pass_fail, item.comments,
                ])
        return output.getvalue()

    def export_all_to_csv(self, forms: list[InspectionForm]) -> str:
        """Export all forms to a single CSV file."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'Form Type', 'System', 'System Tag', 'Section', 'Item ID',
            'Description', 'Check Type', 'Priority', 'Acceptance Criteria',
            'Expected Value', 'Actual Value', 'Pass/Fail', 'Comments',
        ])
        for form in forms:
            for section in form.sections:
                for item in section.check_items:
                    writer.writerow([
                        form.form_type.value, form.system, form.system_tag,
                        section.title, item.id, item.description,
                        item.check_type.value, item.priority.value,
                        item.acceptance_criteria, item.expected_value,
                        item.actual_value, item.pass_fail, item.comments,
                    ])
        return output.getvalue()

    # ── ACC helpers ──────────────────────────────────────────────────

    def _get_response_type(self, check_type: CheckItemType) -> str:
        """Map CheckItemType to ACC response type (comma-delimited)."""
        mapping = {
            CheckItemType.VISUAL: "Yes, No NA",
            CheckItemType.MEASUREMENT: "Text",
            CheckItemType.FUNCTIONAL: "Pass, Fail, N/A",
            CheckItemType.DOCUMENTATION: "Text",
            CheckItemType.VERIFICATION: "Yes, No NA",
        }
        return mapping.get(check_type, "Yes, No NA")

    def _get_non_conforming(self, response_type: str) -> str:
        """Return the non-conforming answer string for a given response type."""
        if response_type == "Yes, No NA":
            return "No"
        if response_type == "Pass, Fail, N/A":
            return "Fail"
        return ""

    def _get_template_name(self, form: InspectionForm, project_number: str = "") -> str:
        """Build ACC-convention template name: {spec}_{level}_{equipment}_{activity}."""
        spec = project_number or form.project or "000000"
        level = FORM_TYPE_LEVEL.get(form.form_type, "L2")
        equipment = (
            form.system_tag
            or form.system.replace(" ", "")[:20]
            or "Equipment"
        )
        activity = FORM_TYPE_ACTIVITY.get(form.form_type, "Inspection")
        return f"{spec}_{level}_{equipment}_{activity}"

    def _get_template_type(self, form: InspectionForm) -> str:
        """Map FormType to ACC Template Type."""
        return FORM_TYPE_TEMPLATE_TYPE.get(form.form_type, "Commissioning")

    # ── ACC Excel export ─────────────────────────────────────────────

    def export_to_acc_excel(
        self,
        forms: list[InspectionForm],
        project_number: str = "",
    ) -> Optional[bytes]:
        """Export forms to ACC (Autodesk Construction Cloud) Excel format.

        Creates a workbook matching the ACC checklist import template with:
        - 22-column structure (A–V)
        - Rows 1-4: Header block (headers, required/optional, descriptions, instructions)
        - Row 5: REMINDER row (red, bold)
        - Row 6+: Data rows (GENERAL section, then form sections + check items)
        - Lookups sheet with valid enumeration values

        Args:
            forms: List of inspection forms to export.
            project_number: Project number for template naming.

        Returns:
            Excel file content as bytes, or None if openpyxl not available.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        wb = Workbook()
        wb.remove(wb.active)

        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        desc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        instr_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        reminder_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        reminder_font = Font(bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        num_cols = len(ACC_HEADERS)

        def _write_header_block(ws):
            """Write the 5-row header block matching the ACC reference template."""
            # Row 1: Column headers
            for col, hdr in enumerate(ACC_HEADERS, 1):
                cell = ws.cell(row=1, column=col, value=hdr)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = wrap_alignment

            # Row 2: Required / Optional labels
            for col, req in enumerate(ACC_ROW2_REQUIRED, 1):
                cell = ws.cell(row=2, column=col, value=req)
                cell.fill = required_fill
                cell.border = thin_border
                cell.alignment = wrap_alignment

            # Row 3: Short descriptions
            for col, desc in enumerate(ACC_ROW3_DESCRIPTIONS, 1):
                cell = ws.cell(row=3, column=col, value=desc)
                cell.fill = desc_fill
                cell.border = thin_border
                cell.alignment = wrap_alignment

            # Row 4: Paste instructions
            for col, instr in enumerate(ACC_ROW4_INSTRUCTIONS, 1):
                cell = ws.cell(row=4, column=col, value=instr)
                cell.fill = instr_fill
                cell.border = thin_border
                cell.alignment = wrap_alignment

            # Row 5: REMINDER row (red background, white bold text)
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=5, column=c)
                cell.fill = reminder_fill
                cell.font = reminder_font
                cell.border = thin_border
            ws.cell(row=5, column=9, value="REMINDER: Don't delete or rename columns. The spreadsheet will not import.\nIt is OK to delete the rows below, including the sample data.")

        def _write_general_section(ws, row, template_name, template_type):
            """Write the GENERAL section with Date, QA/QC Authority, Attendees."""
            general_items = [
                ("GENERAL", "Section", None),
                ("Date", "Date", None),
                ("QA/QC Authority Name", "Text", None),
                ("Inspection Attendees (List Company Names)", "Text", None),
            ]
            for text, resp_type, _ in general_items:
                ws.cell(row=row, column=2, value=template_name)
                ws.cell(row=row, column=3, value=template_type)
                ws.cell(row=row, column=4, value=text)
                ws.cell(row=row, column=6, value=resp_type)
                if resp_type == "Section":
                    for c in range(1, num_cols + 1):
                        ws.cell(row=row, column=c).fill = section_fill
                        ws.cell(row=row, column=c).border = thin_border
                else:
                    for c in range(1, num_cols + 1):
                        ws.cell(row=row, column=c).border = thin_border
                        ws.cell(row=row, column=c).alignment = wrap_alignment
                row += 1
            return row

        def _write_section_row(ws, row, template_name, template_type, section_title):
            """Write a section header row (Response Type = Section)."""
            ws.cell(row=row, column=2, value=template_name)
            ws.cell(row=row, column=3, value=template_type)
            ws.cell(row=row, column=4, value=escape_excel_formula(section_title))
            ws.cell(row=row, column=6, value='Section')
            for c in range(1, num_cols + 1):
                ws.cell(row=row, column=c).fill = section_fill
                ws.cell(row=row, column=c).border = thin_border

        def _write_item_row(ws, row, template_name, template_type, item: CheckItem):
            """Write a single check-item data row."""
            response_type = self._get_response_type(item.check_type)
            non_conforming = self._get_non_conforming(response_type)

            ws.cell(row=row, column=2, value=template_name)
            ws.cell(row=row, column=3, value=template_type)
            ws.cell(row=row, column=4, value=escape_excel_formula(item.description))
            ws.cell(row=row, column=6, value=response_type)
            ws.cell(row=row, column=7, value='TRUE')
            if non_conforming:
                ws.cell(row=row, column=9, value=non_conforming)
            ws.cell(row=row, column=12, value=template_type)
            ws.cell(row=row, column=13, value=template_type)
            ws.cell(row=row, column=19, value='TRUE')

            for c in range(1, num_cols + 1):
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).alignment = wrap_alignment

        def _apply_column_widths(ws):
            for col, w in enumerate(ACC_COLUMN_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(col)].width = w

        DATA_START_ROW = 6  # Data begins after 5 header rows

        # ── Per-form worksheets ──────────────────────────────────────
        for form in forms:
            sheet_name = f"{form.form_type.name}_{form.system_tag}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            template_name = self._get_template_name(form, project_number)
            template_type = self._get_template_type(form)

            _write_header_block(ws)

            row = DATA_START_ROW
            row = _write_general_section(ws, row, template_name, template_type)

            for section in form.sections:
                _write_section_row(ws, row, template_name, template_type, section.title)
                row += 1

                for item in section.check_items:
                    _write_item_row(ws, row, template_name, template_type, item)
                    row += 1

            _apply_column_widths(ws)
            ws.freeze_panes = f'A{DATA_START_ROW}'

        # ── Combined FULL sheet ──────────────────────────────────────
        ws_full = wb.create_sheet(title="FULL", index=0)
        _write_header_block(ws_full)

        row = DATA_START_ROW
        for form in forms:
            template_name = self._get_template_name(form, project_number)
            template_type = self._get_template_type(form)

            row = _write_general_section(ws_full, row, template_name, template_type)

            for section in form.sections:
                _write_section_row(ws_full, row, template_name, template_type, section.title)
                row += 1

                for item in section.check_items:
                    _write_item_row(ws_full, row, template_name, template_type, item)
                    row += 1

        _apply_column_widths(ws_full)
        ws_full.freeze_panes = f'A{DATA_START_ROW}'

        # ── Lookups sheet ────────────────────────────────────────────
        ws_lookup = wb.create_sheet(title="Lookups")
        ws_lookup.cell(row=1, column=1, value="Version")
        ws_lookup.cell(row=1, column=2, value=1)

        ws_lookup.cell(row=3, column=1, value="Template Types")
        for i, tt in enumerate(["Quality", "Safety", "Punch List", "Commissioning"], start=4):
            ws_lookup.cell(row=i, column=1, value=tt)

        ws_lookup.cell(row=9, column=1, value="Response Types")
        response_types = [
            "Section",
            "Yes, No, N/A",
            "Plus, Minus, N/A",
            "True, False, N/A",
            "Pass, Fail, N/A",
            "Text",
            "Date",
        ]
        for i, rt in enumerate(response_types, start=10):
            ws_lookup.cell(row=i, column=1, value=rt)

        ws_lookup.column_dimensions['A'].width = 22
        ws_lookup.column_dimensions['B'].width = 10

        # ── Save ─────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ── Public convenience wrappers ──────────────────────────────────

    def export_to_excel(
        self,
        forms: list[InspectionForm],
        project_number: str = "",
    ) -> Optional[bytes]:
        """Export forms to Excel (ACC format)."""
        return self.export_to_acc_excel(forms, project_number=project_number)

    # Backward-compat alias
    export_to_procore_excel = export_to_acc_excel

    # ── Summary CSV (unchanged) ──────────────────────────────────────

    def export_summary_csv(self, forms: list[InspectionForm]) -> str:
        """Export a summary of all forms to CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            'Form Type', 'System', 'System Tag', 'Total Items',
            'Critical Items', 'High Priority Items',
            'Medium Priority Items', 'Low Priority Items',
        ])
        for form in forms:
            critical = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'CRITICAL')
            high = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'HIGH')
            medium = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'MEDIUM')
            low = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'LOW')
            writer.writerow([
                form.form_type.value, form.system, form.system_tag,
                form.total_items, critical, high, medium, low,
            ])
        return output.getvalue()


def check_excel_support() -> bool:
    """Check if Excel export is available."""
    try:
        import openpyxl
        return True
    except ImportError:
        return False
