"""Export ITC forms to CSV and Excel formats.

Supports Procore-style checklist format for commissioning workflows.
"""

import csv
import io
from typing import Optional

from .models import InspectionForm, FormSection, CheckItem, FormType, CheckItemType


def escape_excel_formula(value: str) -> str:
    """Escape strings that Excel might interpret as formulas.

    Excel treats strings starting with =, -, +, @ as formulas.
    Prefix with a single quote to force text interpretation.
    """
    if not value:
        return value
    if isinstance(value, str) and value and value[0] in ('=', '-', '+', '@'):
        return "'" + value
    return value


class FormExporter:
    """Export inspection forms to various formats."""

    def export_to_csv(self, form: InspectionForm) -> str:
        """Export a single form to CSV format.

        Args:
            form: The inspection form to export

        Returns:
            CSV content as a string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            'Form Type',
            'System',
            'System Tag',
            'Section',
            'Item ID',
            'Description',
            'Check Type',
            'Priority',
            'Acceptance Criteria',
            'Expected Value',
            'Actual Value',
            'Pass/Fail',
            'Comments'
        ])

        # Data rows
        for section in form.sections:
            for item in section.check_items:
                writer.writerow([
                    form.form_type.value,
                    form.system,
                    form.system_tag,
                    section.title,
                    item.id,
                    item.description,
                    item.check_type.value,
                    item.priority.value,
                    item.acceptance_criteria,
                    item.expected_value,
                    item.actual_value,
                    item.pass_fail,
                    item.comments
                ])

        return output.getvalue()

    def export_all_to_csv(self, forms: list[InspectionForm]) -> str:
        """Export all forms to a single CSV file.

        Args:
            forms: List of inspection forms to export

        Returns:
            CSV content as a string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header row
        writer.writerow([
            'Form Type',
            'System',
            'System Tag',
            'Section',
            'Item ID',
            'Description',
            'Check Type',
            'Priority',
            'Acceptance Criteria',
            'Expected Value',
            'Actual Value',
            'Pass/Fail',
            'Comments'
        ])

        # Data rows for all forms
        for form in forms:
            for section in form.sections:
                for item in section.check_items:
                    writer.writerow([
                        form.form_type.value,
                        form.system,
                        form.system_tag,
                        section.title,
                        item.id,
                        item.description,
                        item.check_type.value,
                        item.priority.value,
                        item.acceptance_criteria,
                        item.expected_value,
                        item.actual_value,
                        item.pass_fail,
                        item.comments
                    ])

        return output.getvalue()

    def _get_response_type(self, check_type: CheckItemType) -> str:
        """Map check item type to Procore response type.

        Standard response types from reference:
        - Group Header: Section headers
        - Text: Free text input fields
        - Date: Date input fields
        - Pass, Fail, N/A: Functional tests, verifications
        - Yes, No, N/A: Yes/no questions, confirmations
        - Numeric: Numeric measurements (optional)
        """
        mapping = {
            CheckItemType.VISUAL: "Pass, Fail, N/A",
            CheckItemType.MEASUREMENT: "Text",
            CheckItemType.FUNCTIONAL: "Pass, Fail, N/A",
            CheckItemType.DOCUMENTATION: "Text",
            CheckItemType.VERIFICATION: "Pass, Fail, N/A",
        }
        return mapping.get(check_type, "Yes, No, N/A")

    def _get_checklist_name(self, form: InspectionForm) -> str:
        """Generate Procore-style checklist name."""
        form_type_abbrev = {
            FormType.PFI: "Pre-Functional Inspection",
            FormType.FPT: "Functional Performance Test",
            FormType.IST: "Integrated Systems Test",
            FormType.CXC: "Commissioning Checklist",
        }
        type_name = form_type_abbrev.get(form.form_type, form.form_type.value)
        system_name = form.system[:30] if len(form.system) > 30 else form.system
        return f"[{form.system_tag}]_{system_name}_{type_name}".replace(" ", "_")

    def export_to_procore_excel(self, forms: list[InspectionForm], project_number: str = "") -> Optional[bytes]:
        """Export forms to Procore-style Excel format.

        Creates a workbook matching the Procore checklist import format with:
        - Checklist Name, Permissions, Auto Create Issue, Display Number
        - Item Text, Response Type, Drop-down Answers, Default Answer
        - Answers that create Non-conformance, Default Issue Description
        - Company, Spec Reference, Root Cause Category

        Args:
            forms: List of inspection forms to export
            project_number: Project number for checklist naming

        Returns:
            Excel file content as bytes, or None if openpyxl not available
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return None

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        group_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')

        # Column headers for Procore format (matching reference file)
        headers = [
            '###',
            'Checklist Name',
            'Permissions',
            'Auto Create Issue',
            'Display Number',
            'Item Text',
            'Response Type',
            'Drop-down Answers',
            'Default Answer',
            'Answers that create Non-conformances',
            'Default Issue Description',
            'Company',
            'Spec Reference',
            'Root Cause Category',
            'Root Cause',
            'More Information'
        ]

        # Column widths
        column_widths = [6, 60, 14, 18, 12, 100, 20, 15, 15, 25, 25, 15, 15, 18, 18, 50]

        # Create a worksheet for each form
        for form in forms:
            # Sheet name (max 31 chars)
            sheet_name = f"{form.form_type.name}_{form.system_tag}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            checklist_name = self._get_checklist_name(form)

            # Row 1: Headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = wrap_alignment

            # Row 2: Instructions (optional - can be hidden)
            instructions = [
                '###',
                'Required\n\nName of the checklist',
                'Required\n\nControls access (all, admin)',
                'Optional\n\nAuto-create issues on fail',
                'Optional\n\nItem number',
                'Required\n\nText for the checklist item',
                'Required\n\nResponse type',
                'Optional\n\nCustom dropdown options',
                'Optional\n\nDefault selected answer',
                'Optional\n\nAnswers triggering NC',
                'Optional\n\nDefault issue description',
                'Optional\n\nAssigned company',
                'Optional\n\nSpec section reference',
                'Optional\n\nRoot cause category',
                'Optional\n\nRoot cause details',
                'Optional\n\nAdditional details'
            ]
            for col, instr in enumerate(instructions, 1):
                cell = ws.cell(row=2, column=col, value=instr)
                cell.fill = instruction_fill
                cell.alignment = wrap_alignment
                cell.border = thin_border

            # Row 3: Reminder
            ws.cell(row=3, column=1, value='###')
            ws.cell(row=3, column=2, value="REMINDER: Don't delete or rename columns")

            # Row 4: Headers again (for reference after hiding row 2)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Data rows start at row 5
            row = 5
            display_num = 1

            # Form header row
            ws.cell(row=row, column=2, value=checklist_name)
            ws.cell(row=row, column=3, value='all')
            ws.cell(row=row, column=4, value='True')
            ws.cell(row=row, column=5, value=display_num)
            ws.cell(row=row, column=6, value=form.title)
            ws.cell(row=row, column=7, value='Group Header')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = group_fill
                ws.cell(row=row, column=col).border = thin_border
            row += 1
            display_num += 1

            # Equipment Designation field
            ws.cell(row=row, column=2, value=checklist_name)
            ws.cell(row=row, column=3, value='all')
            ws.cell(row=row, column=4, value='True')
            ws.cell(row=row, column=5, value=display_num)
            ws.cell(row=row, column=6, value='Equipment Designation')
            ws.cell(row=row, column=7, value='Text')
            ws.cell(row=row, column=9, value=form.system_tag)
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
            display_num += 1

            # Commence Date field
            ws.cell(row=row, column=2, value=checklist_name)
            ws.cell(row=row, column=3, value='all')
            ws.cell(row=row, column=4, value='True')
            ws.cell(row=row, column=5, value=display_num)
            ws.cell(row=row, column=6, value='Commence Date')
            ws.cell(row=row, column=7, value='Date')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
            display_num += 1

            # Commissioning Authority field
            ws.cell(row=row, column=2, value=checklist_name)
            ws.cell(row=row, column=3, value='all')
            ws.cell(row=row, column=4, value='True')
            ws.cell(row=row, column=5, value=display_num)
            ws.cell(row=row, column=6, value='Commissioning Authority (Name)')
            ws.cell(row=row, column=7, value='Text')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
            row += 1
            display_num += 1

            # Process each section
            for section in form.sections:
                # Section header
                ws.cell(row=row, column=2, value=checklist_name)
                ws.cell(row=row, column=3, value='all')
                ws.cell(row=row, column=4, value='True')
                ws.cell(row=row, column=5, value=display_num)
                ws.cell(row=row, column=6, value=escape_excel_formula(section.title))
                ws.cell(row=row, column=7, value='Group Header')
                ws.cell(row=row, column=16, value=escape_excel_formula(section.description))
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = group_fill
                    ws.cell(row=row, column=col).border = thin_border
                row += 1
                display_num += 1

                # Section items
                for item in section.check_items:
                    response_type = self._get_response_type(item.check_type)

                    ws.cell(row=row, column=2, value=checklist_name)
                    ws.cell(row=row, column=3, value='all')
                    ws.cell(row=row, column=4, value='True')
                    ws.cell(row=row, column=5, value=display_num)
                    ws.cell(row=row, column=6, value=escape_excel_formula(item.description))
                    ws.cell(row=row, column=7, value=response_type)

                    # Build "More Information" content (column 16)
                    more_info_parts = []
                    if item.expected_value:
                        more_info_parts.append(f"Expected: {item.expected_value}")
                    if item.acceptance_criteria:
                        more_info_parts.append(item.acceptance_criteria)
                    if item.priority.value != "Medium":  # Only note if not default
                        more_info_parts.append(f"Priority: {item.priority.value}")

                    if more_info_parts:
                        ws.cell(row=row, column=16, value=escape_excel_formula(" | ".join(more_info_parts)))

                    # Non-conformance answers
                    if response_type == "Pass, Fail, N/A":
                        ws.cell(row=row, column=10, value="Fail")
                    elif response_type == "Yes, No, N/A":
                        ws.cell(row=row, column=10, value="No")

                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).border = thin_border
                        ws.cell(row=row, column=col).alignment = wrap_alignment

                    row += 1
                    display_num += 1

            # Set column widths
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width

            # Freeze header rows
            ws.freeze_panes = 'A5'

        # Create a combined "FULL" sheet with all forms
        ws_full = wb.create_sheet(title="FULL", index=0)

        # Headers for FULL sheet
        for col, header in enumerate(headers, 1):
            cell = ws_full.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for col, instr in enumerate(instructions, 1):
            cell = ws_full.cell(row=2, column=col, value=instr)
            cell.fill = instruction_fill
            cell.alignment = wrap_alignment
            cell.border = thin_border

        ws_full.cell(row=3, column=1, value='###')
        ws_full.cell(row=3, column=2, value="REMINDER: Don't delete or rename columns")

        for col, header in enumerate(headers, 1):
            cell = ws_full.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Combine all forms into FULL sheet
        row = 5
        display_num = 1

        for form in forms:
            checklist_name = self._get_checklist_name(form)

            # Form header
            ws_full.cell(row=row, column=2, value=checklist_name)
            ws_full.cell(row=row, column=3, value='all')
            ws_full.cell(row=row, column=4, value='True')
            ws_full.cell(row=row, column=5, value=display_num)
            ws_full.cell(row=row, column=6, value=form.title)
            ws_full.cell(row=row, column=7, value='Group Header')
            for col in range(1, len(headers) + 1):
                ws_full.cell(row=row, column=col).fill = group_fill
                ws_full.cell(row=row, column=col).border = thin_border
            row += 1
            display_num += 1

            for section in form.sections:
                # Section header
                ws_full.cell(row=row, column=2, value=checklist_name)
                ws_full.cell(row=row, column=3, value='all')
                ws_full.cell(row=row, column=4, value='True')
                ws_full.cell(row=row, column=5, value=display_num)
                ws_full.cell(row=row, column=6, value=escape_excel_formula(section.title))
                ws_full.cell(row=row, column=7, value='Group Header')
                ws_full.cell(row=row, column=16, value=escape_excel_formula(section.description))
                for col in range(1, len(headers) + 1):
                    ws_full.cell(row=row, column=col).fill = group_fill
                    ws_full.cell(row=row, column=col).border = thin_border
                row += 1
                display_num += 1

                for item in section.check_items:
                    response_type = self._get_response_type(item.check_type)

                    ws_full.cell(row=row, column=2, value=checklist_name)
                    ws_full.cell(row=row, column=3, value='all')
                    ws_full.cell(row=row, column=4, value='True')
                    ws_full.cell(row=row, column=5, value=display_num)
                    ws_full.cell(row=row, column=6, value=escape_excel_formula(item.description))
                    ws_full.cell(row=row, column=7, value=response_type)

                    # Build "More Information" content (column 16)
                    more_info_parts = []
                    if item.expected_value:
                        more_info_parts.append(f"Expected: {item.expected_value}")
                    if item.acceptance_criteria:
                        more_info_parts.append(item.acceptance_criteria)
                    if item.priority.value != "Medium":
                        more_info_parts.append(f"Priority: {item.priority.value}")

                    if more_info_parts:
                        ws_full.cell(row=row, column=16, value=escape_excel_formula(" | ".join(more_info_parts)))

                    if response_type == "Pass, Fail, N/A":
                        ws_full.cell(row=row, column=10, value="Fail")
                    elif response_type == "Yes, No, N/A":
                        ws_full.cell(row=row, column=10, value="No")

                    for col in range(1, len(headers) + 1):
                        ws_full.cell(row=row, column=col).border = thin_border
                        ws_full.cell(row=row, column=col).alignment = wrap_alignment

                    row += 1
                    display_num += 1

        # Set column widths for FULL sheet
        for col, width in enumerate(column_widths, 1):
            ws_full.column_dimensions[get_column_letter(col)].width = width

        ws_full.freeze_panes = 'A5'

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_to_excel(self, forms: list[InspectionForm]) -> Optional[bytes]:
        """Export forms to Excel format (Procore-style).

        This method now uses the Procore format by default.

        Args:
            forms: List of inspection forms to export

        Returns:
            Excel file content as bytes, or None if openpyxl not available
        """
        return self.export_to_procore_excel(forms)

    def export_summary_csv(self, forms: list[InspectionForm]) -> str:
        """Export a summary of all forms to CSV.

        Args:
            forms: List of inspection forms

        Returns:
            CSV content as a string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            'Form Type',
            'System',
            'System Tag',
            'Total Items',
            'Critical Items',
            'High Priority Items',
            'Medium Priority Items',
            'Low Priority Items'
        ])

        # Data
        for form in forms:
            critical = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'CRITICAL')
            high = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'HIGH')
            medium = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'MEDIUM')
            low = sum(1 for s in form.sections for i in s.check_items if i.priority.name == 'LOW')

            writer.writerow([
                form.form_type.value,
                form.system,
                form.system_tag,
                form.total_items,
                critical,
                high,
                medium,
                low
            ])

        return output.getvalue()


def check_excel_support() -> bool:
    """Check if Excel export is available."""
    try:
        import openpyxl
        return True
    except ImportError:
        return False
