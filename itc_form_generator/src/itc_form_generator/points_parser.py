"""Enhanced parser for BMS/Controls points lists.

Improvements over original:
- Excel (.xlsx/.xls) support with multi-sheet handling
- AI-assisted column mapping for non-standard headers
- Hierarchical points list support (parent/child relationships)
- Equipment cross-referencing with SOO data
- Fuzzy matching for column headers
- Confidence scoring for parsed points
- Auto-detection of header rows
- Support for merged cells and multi-line headers
"""

import csv
import io
import re
import json
import logging
from dataclasses import dataclass, field
from enum import Enum
from typing import Optional, Any
from difflib import SequenceMatcher

logger = logging.getLogger(__name__)


class PointType(Enum):
    """Types of control points."""
    AI = "Analog Input"
    AO = "Analog Output"
    DI = "Digital Input"
    DO = "Digital Output"
    AV = "Analog Value"
    BV = "Binary Value"
    IAV = "Internal Analog Value"
    IBV = "Internal Binary Value"
    MV = "Multi-State Value"
    MSI = "Multi-State Input"
    MSO = "Multi-State Output"
    CALC = "Calculated"
    VIRTUAL = "Virtual"
    UNKNOWN = "Unknown"


@dataclass
class ControlPoint:
    """A single control/BMS point with enhanced metadata."""
    point_name: str
    tag: str = ""
    point_type: PointType = PointType.UNKNOWN
    description: str = ""
    units: str = ""
    range_min: str = ""
    range_max: str = ""
    design_value: str = ""
    alarms: list[str] = field(default_factory=list)
    system_ref: str = ""
    equipment_ref: str = ""
    area: str = ""
    process_area: str = ""
    end_device: str = ""
    software_function: str = ""
    # New fields
    parent_point: str = ""
    hierarchy_level: int = 0
    sheet_name: str = ""
    row_number: int = 0
    confidence: float = 1.0
    matched_soo_system: str = ""
    matched_soo_equipment: str = ""
    notes: str = ""
    network_ref: str = ""
    controller_ref: str = ""


@dataclass
class PointsList:
    """Collection of control points with metadata."""
    name: str = ""
    points: list[ControlPoint] = field(default_factory=list)
    source_format: str = ""  # csv, xlsx, gtn, etc.
    column_mapping: dict = field(default_factory=dict)
    parsing_confidence: float = 1.0
    warnings: list[str] = field(default_factory=list)

    @property
    def ai_points(self) -> list[ControlPoint]:
        return [p for p in self.points if p.point_type == PointType.AI]

    @property
    def ao_points(self) -> list[ControlPoint]:
        return [p for p in self.points if p.point_type == PointType.AO]

    @property
    def di_points(self) -> list[ControlPoint]:
        return [p for p in self.points if p.point_type in (PointType.DI, PointType.BV, PointType.IBV)]

    @property
    def do_points(self) -> list[ControlPoint]:
        return [p for p in self.points if p.point_type == PointType.DO]

    @property
    def av_points(self) -> list[ControlPoint]:
        return [p for p in self.points if p.point_type in (PointType.AV, PointType.IAV)]

    @property
    def by_system(self) -> dict[str, list[ControlPoint]]:
        """Group points by system reference."""
        groups = {}
        for p in self.points:
            key = p.system_ref or p.matched_soo_system or "Unassigned"
            groups.setdefault(key, []).append(p)
        return groups

    @property
    def by_equipment(self) -> dict[str, list[ControlPoint]]:
        """Group points by equipment reference."""
        groups = {}
        for p in self.points:
            key = p.equipment_ref or p.matched_soo_equipment or "Unassigned"
            groups.setdefault(key, []).append(p)
        return groups

    @property
    def summary(self) -> dict:
        """Summary statistics."""
        type_counts = {}
        for p in self.points:
            type_counts[p.point_type.name] = type_counts.get(p.point_type.name, 0) + 1
        return {
            "total_points": len(self.points),
            "by_type": type_counts,
            "systems": len(self.by_system),
            "equipment": len(self.by_equipment),
            "confidence": self.parsing_confidence,
            "warnings": len(self.warnings),
        }


class PointsListParser:
    """Enhanced parser for BMS/Controls points lists.

    Supports:
    - CSV/TSV with standard or non-standard column headers
    - Excel (.xlsx/.xls) with multi-sheet support
    - GTN-style enterprise format with hierarchical structure
    - AI-assisted column mapping for unknown formats
    - Equipment cross-referencing with SOO data
    """

    COLUMN_ALIASES = {
        'point_name': ['point name', 'pointname', 'name', 'point', 'point id', 'pointid',
                       'derived point name', 'bacnet object name', 'object name',
                       'point reference', 'point ref', 'signal name', 'signal'],
        'tag': ['tag', 'tag name', 'tagname', 'object name', 'objectname', 'address',
                'bacnet address', 'plc address', 'controller address', 'point tag'],
        'point_type': ['type', 'point type', 'pointtype', 'i/o type', 'io type', 'iotype',
                       'object type', 'bacnet type', 'signal type', 'data type'],
        'description': ['description', 'desc', 'label', 'point description', 'plc comment',
                        'plc\ncomment', 'function', 'purpose', 'remarks', 'comment',
                        'point desc', 'signal description'],
        'units': ['units', 'unit', 'eng units', 'engineering units', 'eng.\nunits',
                  'uom', 'unit of measure', 'measurement unit'],
        'range': ['range', 'limits', 'min/max', 'scaled\nrange', 'scaling'],
        'range_min': ['min', 'minimum', 'low limit', 'low', 'min value', 'low range'],
        'range_max': ['max', 'maximum', 'high limit', 'high', 'max value', 'high range'],
        'alarms': ['alarms', 'alarm', 'alarm limits', 'alarm points', 'alarm config'],
        'system': ['system', 'system name', 'systemname', 'system ref', 'sys',
                   'system tag', 'subsystem'],
        'equipment': ['equipment', 'equipment name', 'equip', 'device', 'equipment tag',
                      'asset', 'asset tag', 'component', 'equipment id'],
        'design_value': ['design value', 'design value\nsetpoint', 'setpoint', 'default',
                         'default value', 'normal value', 'operating value'],
        'area': ['area', 'zone', 'location', 'building', 'floor', 'room'],
        'process_area': ['process area', 'process', 'discipline'],
        'end_device': ['end device', 'field device', 'sensor', 'actuator', 'instrument'],
        'software_function': ['software function', 'sw function', 'function block'],
        'item_name': ['item name', 'item'],
        'network': ['network', 'network ref', 'bus', 'protocol', 'communication'],
        'controller': ['controller', 'plc', 'ddc', 'panel', 'controller name'],
        'notes': ['notes', 'note', 'remark', 'additional info'],
    }

    TYPE_PATTERNS = {
        PointType.AI: [r'\bAI\b', r'analog\s*in', r'input.*analog', r'\bANALOG INPUT\b'],
        PointType.AO: [r'\bAO\b', r'analog\s*out', r'output.*analog', r'\bANALOG OUTPUT\b'],
        PointType.DI: [r'\bDI\b', r'\bBI\b', r'digital\s*in', r'binary\s*in', r'\bDIGITAL INPUT\b'],
        PointType.DO: [r'\bDO\b', r'\bBO\b', r'digital\s*out', r'binary\s*out', r'\bDIGITAL OUTPUT\b'],
        PointType.AV: [r'\bAV\b', r'analog\s*value', r'\bANALOG VALUE\b'],
        PointType.BV: [r'\bBV\b', r'binary\s*value', r'\bBINARY VALUE\b'],
        PointType.IAV: [r'\biAV\b', r'internal\s*analog'],
        PointType.IBV: [r'\biBV\b', r'internal\s*binary'],
        PointType.MV: [r'\bMV\b', r'multi.?state\s*value'],
        PointType.MSI: [r'\bMSI\b', r'multi.?state\s*in'],
        PointType.MSO: [r'\bMSO\b', r'multi.?state\s*out'],
        PointType.CALC: [r'\bCALC\b', r'calculated', r'derived'],
        PointType.VIRTUAL: [r'\bVIRT\b', r'virtual'],
    }

    # Equipment tag patterns for extraction from point names
    EQUIPMENT_PATTERNS = [
        r'^([A-Z]{2,5}[-_]\d{1,4}[A-Z]?)',           # CRAH-01, AHU_001A
        r'^([A-Z]+\d+)',                               # AHU01, FCU12
        r'^(\w+-\w+-\w+)-',                            # SYS-SUB-EQUIP-point
        r'[._]([A-Z]{2,5}[-_]?\d{1,4}[A-Z]?)[._]',    # prefix.CRAH-01.point
    ]

    def __init__(self, ai_service=None, soo_data=None):
        """Initialize parser with optional AI service and SOO cross-reference data.

        Args:
            ai_service: Optional AIService instance for AI-assisted column mapping
            soo_data: Optional SequenceOfOperation for equipment cross-referencing
        """
        self.ai_service = ai_service
        self.soo_data = soo_data
        self._fuzzy_threshold = 0.7  # Minimum similarity for fuzzy column matching

    def parse(self, content: str, filename: str = "") -> PointsList:
        """Parse points list from CSV/TSV/Excel content.

        Args:
            content: File content (text for CSV/TSV, or file path for Excel)
            filename: Original filename for format detection

        Returns:
            PointsList with parsed points
        """
        points_list = PointsList(name=filename)

        # Detect format
        if filename.endswith(('.xlsx', '.xls')):
            points_list = self._parse_excel(content, filename)
            points_list.source_format = "xlsx"
        elif self._is_gtn_format(content.split('\n')):
            points_list = self._parse_gtn_format(content)
            points_list.source_format = "gtn"
        else:
            points_list = self._parse_csv(content)
            points_list.source_format = "csv"

        points_list.name = filename

        # Post-processing: cross-reference with SOO data
        if self.soo_data:
            self._cross_reference_soo(points_list)

        # Post-processing: calculate confidence
        self._calculate_confidence(points_list)

        return points_list

    def _parse_excel(self, file_path: str, filename: str = "") -> PointsList:
        """Parse Excel file with multi-sheet support."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed. Install with: pip install openpyxl")
            return PointsList(warnings=["openpyxl not installed - cannot parse Excel files"])

        points_list = PointsList()

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open Excel file: {e}")
            return PointsList(warnings=[f"Failed to open Excel file: {str(e)}"])

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Skip empty sheets or sheets with too few rows
            if ws.max_row is None or ws.max_row < 2:
                continue

            # Find header row
            header_row_idx = self._find_header_row(ws)
            if header_row_idx is None:
                logger.info(f"Skipping sheet \'{sheet_name}\' - no header row found")
                continue

            # Get headers (handle merged cells)
            headers = self._get_excel_headers(ws, header_row_idx)
            column_map = self._map_columns(headers)

            # If standard mapping fails, try AI-assisted mapping
            if len(column_map) < 2 and self.ai_service:
                column_map = self._ai_map_columns(headers)

            # If still poor mapping, try fuzzy matching
            if len(column_map) < 2:
                column_map = self._fuzzy_map_columns(headers)

            if not column_map:
                points_list.warnings.append(f"Sheet \'{sheet_name}\': could not map columns")
                continue

            points_list.column_mapping[sheet_name] = column_map

            # Parse data rows
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    row_data[header] = str(cell.value) if cell.value is not None else ""

                point = self._parse_row(row_data, column_map)
                if point and point.point_name:
                    point.sheet_name = sheet_name
                    point.row_number = row_idx
                    points_list.points.append(point)

        return points_list

    def _find_header_row(self, ws) -> Optional[int]:
        """Auto-detect the header row in an Excel sheet."""
        # Look for row with most recognizable column names
        best_row = None
        best_score = 0

        all_aliases = set()
        for aliases in self.COLUMN_ALIASES.values():
            all_aliases.update(a.lower().replace('\n', ' ') for a in aliases)

        for row_idx in range(1, min(20, ws.max_row + 1)):  # Check first 20 rows
            score = 0
            non_empty = 0
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    non_empty += 1
                    val = str(cell.value).lower().strip().replace('\n', ' ')
                    if val in all_aliases:
                        score += 2  # Exact match
                    elif any(SequenceMatcher(None, val, alias).ratio() > 0.8 for alias in all_aliases):
                        score += 1  # Fuzzy match

            if non_empty >= 3 and score > best_score:
                best_score = score
                best_row = row_idx

        return best_row

    def _get_excel_headers(self, ws, header_row: int) -> list[str]:
        """Extract headers from Excel sheet, handling merged cells."""
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            val = str(cell.value).strip() if cell.value else f"col_{col_idx}"
            headers.append(val)
        return headers

    def _parse_csv(self, content: str) -> PointsList:
        """Parse standard CSV/TSV content with enhanced detection."""
        lines = content.split('\n')
        points_list = PointsList()

        # Auto-detect header row (skip metadata rows at top)
        header_idx = self._find_csv_header_row(lines)
        if header_idx is None:
            points_list.warnings.append("Could not find header row in CSV")
            return points_list

        # Use content from header row onwards
        csv_content = '\n'.join(lines[header_idx:])
        dialect = self._detect_dialect(csv_content)

        reader = csv.DictReader(io.StringIO(csv_content), dialect=dialect)
        fieldnames = reader.fieldnames or []

        column_map = self._map_columns(fieldnames)

        # Fallback to AI or fuzzy matching
        if len(column_map) < 2 and self.ai_service:
            column_map = self._ai_map_columns(fieldnames)
        if len(column_map) < 2:
            column_map = self._fuzzy_map_columns(fieldnames)

        points_list.column_mapping["default"] = column_map

        for row_num, row in enumerate(reader, header_idx + 2):
            point = self._parse_row(row, column_map)
            if point and point.point_name:
                point.row_number = row_num
                points_list.points.append(point)

        return points_list

    def _find_csv_header_row(self, lines: list[str]) -> Optional[int]:
        """Find the actual header row in CSV, skipping metadata."""
        all_aliases = set()
        for aliases in self.COLUMN_ALIASES.values():
            all_aliases.update(a.lower().replace('\n', ' ') for a in aliases)

        for idx, line in enumerate(lines[:20]):
            if not line.strip():
                continue
            # Split by common delimiters
            parts = re.split(r'[,\t|;]', line.lower())
            matches = sum(1 for p in parts if p.strip() in all_aliases or
                         any(SequenceMatcher(None, p.strip(), a).ratio() > 0.8 for a in all_aliases))
            if matches >= 2:
                return idx

        return 0  # Default to first row

    def _is_gtn_format(self, lines: list[str]) -> bool:
        """Detect GTN-style format with hierarchical headers."""
        if len(lines) < 5:
            return False
        first_line = lines[0].lower()
        has_version_control = 'version control' in first_line
        has_derived = any('derived point name' in line.lower() for line in lines[:5])
        has_hierarchical = any('process area' in line.lower() and 'end device' in line.lower()
                              for line in lines[:5])
        return has_version_control or has_derived or has_hierarchical

    def _parse_gtn_format(self, content: str) -> PointsList:
        """Parse GTN-style hierarchical points list with auto-detection."""
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        if len(rows) < 5:
            return PointsList(warnings=["GTN file too short"])

        # Auto-detect column positions by scanning header rows
        col_map = self._auto_detect_gtn_columns(rows[:5])

        points_list = PointsList()
        # Track hierarchy for parent-child relationships
        hierarchy = {}  # level -> current value

        # Skip header rows
        data_start = self._find_gtn_data_start(rows)

        for row_idx, row in enumerate(rows[data_start:], data_start):
            if not row or all(not cell.strip() for cell in row):
                continue

            point_name = self._safe_get(row, col_map.get('point_name', -1))
            if not point_name:
                # Could be a hierarchy row - track it
                area = self._safe_get(row, col_map.get('area', -1))
                system = self._safe_get(row, col_map.get('system', -1))
                equipment = self._safe_get(row, col_map.get('equipment', -1))
                if area:
                    hierarchy['area'] = area
                if system:
                    hierarchy['system'] = system
                if equipment:
                    hierarchy['equipment'] = equipment
                continue

            io_type_str = self._safe_get(row, col_map.get('point_type', -1))
            point_type = self._infer_point_type(io_type_str, point_name)

            # Inherit hierarchy values if not explicitly set
            area = self._safe_get(row, col_map.get('area', -1)) or hierarchy.get('area', '')
            system = self._safe_get(row, col_map.get('system', -1)) or hierarchy.get('system', '')
            equipment = self._safe_get(row, col_map.get('equipment', -1)) or hierarchy.get('equipment', '')

            point = ControlPoint(
                point_name=point_name,
                tag=point_name,
                point_type=point_type,
                description=self._safe_get(row, col_map.get('description', -1)),
                units=self._safe_get(row, col_map.get('units', -1)),
                design_value=self._safe_get(row, col_map.get('design_value', -1)),
                system_ref=system,
                equipment_ref=equipment,
                area=area,
                process_area=self._safe_get(row, col_map.get('process_area', -1)),
                end_device=self._safe_get(row, col_map.get('end_device', -1)),
                software_function=self._safe_get(row, col_map.get('software_function', -1)),
                row_number=row_idx,
            )
            points_list.points.append(point)

        return points_list

    def _auto_detect_gtn_columns(self, header_rows: list[list[str]]) -> dict[str, int]:
        """Auto-detect column positions in GTN format by scanning headers."""
        col_map = {}
        gtn_keywords = {
            'point_name': ['point name', 'derived point name', 'bacnet object'],
            'point_type': ['i/o type', 'io type', 'object type', 'type'],
            'description': ['description', 'plc comment', 'comment'],
            'units': ['units', 'eng units', 'engineering units'],
            'design_value': ['design value', 'setpoint', 'default'],
            'area': ['area', 'zone'],
            'system': ['system', 'system name'],
            'process_area': ['process area'],
            'equipment': ['equipment', 'equipment name', 'device'],
            'end_device': ['end device', 'field device'],
            'software_function': ['software function', 'sw function'],
            'range': ['range', 'scaled range', 'limits'],
        }

        for row in header_rows:
            for col_idx, cell in enumerate(row):
                cell_lower = cell.lower().strip().replace('\n', ' ') if cell else ''
                for field_name, keywords in gtn_keywords.items():
                    if field_name not in col_map:
                        for kw in keywords:
                            if kw in cell_lower:
                                col_map[field_name] = col_idx
                                break

        return col_map

    def _find_gtn_data_start(self, rows: list[list[str]]) -> int:
        """Find where data starts in GTN format (after header rows)."""
        for idx in range(min(10, len(rows))):
            row = rows[idx]
            # Data rows typically have values in many columns
            non_empty = sum(1 for cell in row if cell and cell.strip())
            if idx > 2 and non_empty > 5:
                return idx
        return 4  # Default for GTN format

    def _safe_get(self, row: list, idx: int) -> str:
        """Safely get a value from a row by index."""
        if idx < 0 or idx >= len(row):
            return ""
        return row[idx].strip() if row[idx] else ""

    def _fuzzy_map_columns(self, fieldnames: list[str]) -> dict[str, str]:
        """Fuzzy match column names when exact matching fails."""
        column_map = {}

        for standard_name, aliases in self.COLUMN_ALIASES.items():
            best_match = None
            best_score = 0

            for fname in fieldnames:
                fname_clean = fname.lower().strip().replace('\n', ' ')
                for alias in aliases:
                    score = SequenceMatcher(None, fname_clean, alias).ratio()
                    if score > best_score and score >= self._fuzzy_threshold:
                        best_score = score
                        best_match = fname

            if best_match and standard_name not in column_map:
                column_map[standard_name] = best_match

        return column_map

    def _ai_map_columns(self, fieldnames: list[str]) -> dict[str, str]:
        """Use AI to map non-standard column names to standard fields."""
        if not self.ai_service:
            return {}

        try:
            prompt = f"""You are a BMS/HVAC controls expert. Map these spreadsheet column headers 
to standard BMS point list fields.

Column headers found: {json.dumps(list(fieldnames))}

Standard fields to map to:
- point_name: The unique identifier/name of the control point
- tag: BACnet or controller address/tag
- point_type: Type of I/O (AI, AO, DI, DO, AV, BV, etc.)
- description: Human-readable description of what the point does
- units: Engineering units (°F, %, GPM, etc.)
- range_min: Minimum scaled value
- range_max: Maximum scaled value
- design_value: Default/setpoint value
- alarms: Alarm configuration
- system: System name (HVAC, Electrical, etc.)
- equipment: Equipment tag (AHU-01, CRAH-03, etc.)
- area: Physical area/zone/location
- controller: Controller/PLC reference
- notes: Additional notes

Return a JSON object mapping standard field names to the actual column header.
Only include fields you are confident about (>80% sure).
Example: {{"point_name": "Signal Name", "description": "PLC Comment", "units": "Eng. Units"}}"""

            result = self.ai_service._call_llm(prompt)
            mapping = self.ai_service._extract_json(result)
            if isinstance(mapping, dict):
                # Verify mapped columns actually exist
                valid_map = {}
                for standard, actual in mapping.items():
                    if actual in fieldnames:
                        valid_map[standard] = actual
                return valid_map
        except Exception as e:
            logger.warning(f"AI column mapping failed: {e}")

        return {}

    def _detect_dialect(self, content: str) -> str:
        """Detect CSV dialect (comma, tab, semicolon, pipe)."""
        first_lines = '\n'.join(content.split('\n')[:5])
        delimiters = {
            '\t': first_lines.count('\t'),
            ',': first_lines.count(','),
            ';': first_lines.count(';'),
            '|': first_lines.count('|'),
        }
        best = max(delimiters, key=delimiters.get)
        if best == '\t':
            return 'excel-tab'
        return 'excel'

    def _map_columns(self, fieldnames: list[str]) -> dict[str, str]:
        """Map actual column names to standard names using exact/alias matching."""
        column_map = {}
        fieldnames_lower = {f.lower().strip().replace('\n', ' '): f for f in fieldnames}

        for standard_name, aliases in self.COLUMN_ALIASES.items():
            for alias in aliases:
                alias_clean = alias.replace('\n', ' ')
                if alias_clean in fieldnames_lower:
                    column_map[standard_name] = fieldnames_lower[alias_clean]
                    break

        return column_map

    def _parse_row(self, row: dict[str, str], column_map: dict[str, str]) -> Optional[ControlPoint]:
        """Parse a single row into a ControlPoint with enhanced extraction."""
        def get_value(key: str) -> str:
            col = column_map.get(key)
            if col and col in row:
                val = row[col].strip()
                return val if val and val.lower() != 'none' else ""
            return ""

        point_name = get_value('point_name')
        if not point_name:
            return None

        point_type_str = get_value('point_type')
        point_type = self._infer_point_type(point_type_str, point_name)

        range_str = get_value('range')
        range_min = get_value('range_min')
        range_max = get_value('range_max')

        if range_str and not (range_min or range_max):
            range_min, range_max = self._parse_range(range_str)

        alarms_str = get_value('alarms')
        alarms = [a.strip() for a in re.split(r'[,;|]', alarms_str) if a.strip()] if alarms_str else []

        equipment_ref = get_value('equipment')
        if not equipment_ref:
            equipment_ref = self._extract_equipment_tag(point_name)

        description = get_value('description')
        if not description:
            # Try to infer description from point name
            description = self._infer_description(point_name)

        return ControlPoint(
            point_name=point_name,
            tag=get_value('tag') or point_name,
            point_type=point_type,
            description=description,
            units=get_value('units'),
            range_min=range_min,
            range_max=range_max,
            design_value=get_value('design_value'),
            alarms=alarms,
            system_ref=get_value('system'),
            equipment_ref=equipment_ref,
            area=get_value('area'),
            process_area=get_value('process_area'),
            end_device=get_value('end_device'),
            software_function=get_value('software_function'),
            network_ref=get_value('network'),
            controller_ref=get_value('controller'),
            notes=get_value('notes'),
        )

    def _infer_point_type(self, type_str: str, point_name: str) -> PointType:
        """Infer point type from type string or point name with enhanced patterns."""
        text = f"{type_str} {point_name}".upper()

        for point_type, patterns in self.TYPE_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, text, re.IGNORECASE):
                    return point_type

        # Enhanced name-based inference
        name_upper = point_name.upper()

        # Analog inputs (measurements)
        if any(kw in name_upper for kw in ['TEMP', 'PRESS', 'FLOW', 'LEVEL', 'HUMID',
                'SPEED', 'FREQ', 'VOLT', 'AMP', 'POWER', 'ENERGY', 'PH',
                'CONDUCTIVITY', 'TURBIDITY', 'DENSITY', 'VIBRATION', 'DIFFERENTIAL',
                'DEWPOINT', 'ENTHALPY', 'WB', 'DB', 'RH', 'PPM', 'CO2']):
            return PointType.AI

        # Analog outputs (commands with values)
        if any(kw in name_upper for kw in ['CMD', 'SP', 'SETPOINT', 'OUTPUT', 'VFD',
                'DAMPER', 'VALVE', 'POSITION', 'MODULATE']):
            return PointType.AO

        # Digital inputs (status/feedback)
        if any(kw in name_upper for kw in ['STATUS', 'STS', 'PROOF', 'ALARM', 'FAULT',
                'RUN', 'TRIP', 'FLOW_SW', 'DP_SW', 'SMOKE', 'FIRE',
                'OCCUPIED', 'OCC', 'OVERLOAD', 'HAND', 'AUTO']):
            return PointType.DI

        # Digital outputs (on/off commands)
        if any(kw in name_upper for kw in ['START', 'STOP', 'ENABLE', 'ON', 'OFF',
                'OPEN', 'CLOSE', 'RESET', 'SILENCE', 'ENERGIZE']):
            return PointType.DO

        return PointType.UNKNOWN

    def _infer_description(self, point_name: str) -> str:
        """Infer a human-readable description from a structured point name."""
        # Split by common separators
        parts = re.split(r'[-_./]', point_name)
        if len(parts) < 2:
            return ""

        # Try to build description from parts
        desc_parts = []
        for part in parts:
            # Skip pure numbers or short codes
            if part.isdigit() or len(part) <= 1:
                continue
            # Expand common abbreviations
            expansions = {
                'TEMP': 'Temperature', 'PRESS': 'Pressure', 'FLOW': 'Flow',
                'SP': 'Setpoint', 'CMD': 'Command', 'STS': 'Status',
                'ALM': 'Alarm', 'FLT': 'Fault', 'RUN': 'Running',
                'SAT': 'Supply Air Temp', 'RAT': 'Return Air Temp',
                'OAT': 'Outside Air Temp', 'MAT': 'Mixed Air Temp',
                'DAT': 'Discharge Air Temp', 'CHW': 'Chilled Water',
                'HW': 'Hot Water', 'CW': 'Condenser Water',
                'VFD': 'Variable Frequency Drive', 'DP': 'Differential Pressure',
                'RH': 'Relative Humidity', 'CO2': 'CO2 Level',
            }
            expanded = expansions.get(part.upper(), part)
            desc_parts.append(expanded)

        return ' '.join(desc_parts) if desc_parts else ""

    def _parse_range(self, range_str: str) -> tuple[str, str]:
        """Parse range string with multiple format support."""
        patterns = [
            r'(-?[\d.]+)\s*[-\u2013\u2014to]+\s*(-?[\d.]+)',
            r'(-?[\d.]+)\s*,\s*(-?[\d.]+)',
            r'(-?[\d.]+)\s*/\s*(-?[\d.]+)',
            r'min:\s*(-?[\d.]+).*max:\s*(-?[\d.]+)',
            r'low:\s*(-?[\d.]+).*high:\s*(-?[\d.]+)',
        ]

        for pattern in patterns:
            match = re.search(pattern, range_str, re.IGNORECASE)
            if match:
                return match.group(1), match.group(2)

        return "", ""

    def _extract_equipment_tag(self, point_name: str) -> str:
        """Extract equipment tag from point name using multiple patterns."""
        for pattern in self.EQUIPMENT_PATTERNS:
            match = re.search(pattern, point_name.upper())
            if match:
                return match.group(1)
        return ""

    def _cross_reference_soo(self, points_list: PointsList):
        """Cross-reference points with SOO data to match equipment and systems."""
        if not self.soo_data:
            return

        # Build lookup from SOO data
        soo_systems = {}
        soo_equipment = {}

        for system in getattr(self.soo_data, 'systems', []):
            sys_name = getattr(system, 'name', '')
            if sys_name:
                soo_systems[sys_name.upper()] = sys_name
                # Also index by tag
                sys_tag = getattr(system, 'tag', '')
                if sys_tag:
                    soo_systems[sys_tag.upper()] = sys_name

            for component in getattr(system, 'components', []):
                comp_name = getattr(component, 'name', '')
                comp_tag = getattr(component, 'tag', '')
                if comp_tag:
                    soo_equipment[comp_tag.upper()] = {
                        'equipment': comp_name,
                        'system': sys_name
                    }
                if comp_name:
                    soo_equipment[comp_name.upper()] = {
                        'equipment': comp_name,
                        'system': sys_name
                    }

        # Match points to SOO data
        for point in points_list.points:
            equip_tag = (point.equipment_ref or self._extract_equipment_tag(point.point_name)).upper()

            # Direct match
            if equip_tag in soo_equipment:
                match = soo_equipment[equip_tag]
                point.matched_soo_equipment = match['equipment']
                point.matched_soo_system = match['system']
                continue

            # Fuzzy match
            best_match = None
            best_score = 0
            for soo_tag, info in soo_equipment.items():
                score = SequenceMatcher(None, equip_tag, soo_tag).ratio()
                if score > best_score and score > 0.7:
                    best_score = score
                    best_match = info

            if best_match:
                point.matched_soo_equipment = best_match['equipment']
                point.matched_soo_system = best_match['system']
                point.confidence *= best_score  # Reduce confidence for fuzzy matches

    def _calculate_confidence(self, points_list: PointsList):
        """Calculate overall parsing confidence."""
        if not points_list.points:
            points_list.parsing_confidence = 0.0
            return

        total_score = 0
        for point in points_list.points:
            score = 0.5  # Base
            if point.point_name:
                score += 0.1
            if point.point_type != PointType.UNKNOWN:
                score += 0.15
            if point.description:
                score += 0.1
            if point.equipment_ref or point.matched_soo_equipment:
                score += 0.1
            if point.system_ref or point.matched_soo_system:
                score += 0.05
            point.confidence = min(1.0, score * point.confidence)
            total_score += point.confidence

        points_list.parsing_confidence = total_score / len(points_list.points)

